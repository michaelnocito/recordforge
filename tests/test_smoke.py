"""Phase A smoke tests."""

import random
from decimal import Decimal
from pathlib import Path

import pytest

from recordforge.core.models import DocumentData, GeneratedDoc
from recordforge.core.seed import get_rng, set_seed


# --- Generator smoke tests ---

def _fresh_rng() -> random.Random:
    return random.Random(12345)


def test_invoice_build():
    from recordforge.generators.documents.invoice import build
    data = build(_fresh_rng())
    assert isinstance(data, DocumentData)
    assert data.doc_type == "invoice"
    assert data.doc_number.startswith("INV-")
    assert 2 <= len(data.line_items) <= 4


def test_purchase_order_build():
    from recordforge.generators.documents.purchase_order import build
    data = build(_fresh_rng())
    assert data.doc_type == "purchase_order"
    assert data.doc_number.startswith("PO-")
    assert 3 <= len(data.line_items) <= 6


def test_intake_form_build():
    from recordforge.generators.documents.intake_form import build
    data = build(_fresh_rng())
    assert data.doc_type == "intake_form"
    assert len(data.line_items) == 0
    assert data.notes


def test_sop_build():
    from recordforge.generators.documents.sop import build
    data = build(_fresh_rng())
    assert data.doc_type == "sop"
    assert len(data.line_items) == 0


def test_contract_build():
    from recordforge.generators.documents.contract import build
    data = build(_fresh_rng())
    assert data.doc_type == "contract"
    assert len(data.line_items) == 0
    assert "SERVICES AGREEMENT" in data.notes


def test_offer_letter_build():
    from recordforge.generators.documents.offer_letter import build
    data = build(_fresh_rng())
    assert data.doc_type == "offer_letter"
    assert len(data.line_items) == 0
    assert "Base Salary" in data.notes


# --- Line item math ---

def test_line_item_totals_are_computed():
    from recordforge.generators.documents.invoice import build
    data = build(_fresh_rng())
    for item in data.line_items:
        assert item.total == Decimal(item.quantity) * item.unit_price
    assert data.subtotal == sum(i.total for i in data.line_items)
    assert data.total_due == data.subtotal + data.tax


# --- Date ordering ---

def test_invoice_due_date_after_doc_date():
    from datetime import datetime
    from recordforge.generators.documents.invoice import build
    data = build(_fresh_rng())
    fmt = "%B %d, %Y"
    doc = datetime.strptime(data.doc_date, fmt)
    due = datetime.strptime(data.due_date, fmt)
    assert due > doc


# --- Data generator smoke tests ---

def test_customers_build_rows():
    from recordforge.generators.data.customers import build_rows
    rows = build_rows(_fresh_rng(), count=10)
    assert len(rows) == 10
    assert "customer_id" in rows[0]
    assert rows[0]["customer_id"] == "CUST-1000"


def test_vendors_build_rows():
    from recordforge.generators.data.vendors import build_rows
    rows = build_rows(_fresh_rng(), count=5)
    assert len(rows) == 5
    assert "vendor_id" in rows[0]


def test_transactions_build_rows():
    from recordforge.generators.data.transactions import build_rows
    rows = build_rows(_fresh_rng(), count=5)
    assert all(r["currency"] == "USD" for r in rows)


def test_employees_build_rows():
    from recordforge.generators.data.employees import build_rows
    rows = build_rows(_fresh_rng(), count=5)
    assert "employee_id" in rows[0]


def test_inventory_build_rows():
    from recordforge.generators.data.inventory import build_rows
    rows = build_rows(_fresh_rng(), count=5)
    assert "sku" in rows[0]


def test_messy_build_rows():
    from recordforge.generators.data.messy import build_rows
    rows = build_rows(_fresh_rng(), count=20)
    # Must contain at least some None values (dirty by design)
    all_values = [v for row in rows for v in row.values()]
    assert None in all_values


# --- Renderer smoke tests ---

def test_pdf_renderer_produces_nonempty_file(tmp_path: Path):
    from recordforge.generators.documents.invoice import build
    from recordforge.renderers.pdf import render
    data = build(_fresh_rng())
    doc = render(data, tmp_path)
    assert isinstance(doc, GeneratedDoc)
    assert doc.path.exists()
    assert doc.path.stat().st_size > 0


def test_docx_renderer_produces_nonempty_file(tmp_path: Path):
    from recordforge.generators.documents.invoice import build
    from recordforge.renderers.docx import render
    data = build(_fresh_rng())
    doc = render(data, tmp_path)
    assert doc.path.exists()
    assert doc.path.stat().st_size > 0


def test_html_renderer_produces_nonempty_file(tmp_path: Path):
    from recordforge.generators.documents.invoice import build
    from recordforge.renderers.html import render
    data = build(_fresh_rng())
    doc = render(data, tmp_path)
    assert doc.path.exists()
    content = doc.path.read_text(encoding="utf-8")
    assert "SAMPLE" in content
    assert data.doc_number in content


def test_xlsx_renderer_produces_nonempty_file(tmp_path: Path):
    from recordforge.generators.data.customers import build_rows
    from recordforge.renderers.xlsx import render
    rows = build_rows(_fresh_rng(), count=10)
    doc = render("customers", rows, tmp_path)
    assert doc.path.exists()
    assert doc.path.stat().st_size > 0


# --- Watermark test ---

def test_pdf_contains_sample_watermark(tmp_path: Path):
    """Generated PDF contains watermark indicators: Helvetica-Bold font and 0.15 alpha."""
    from recordforge.generators.documents.invoice import build
    from recordforge.renderers.pdf import render
    data = build(_fresh_rng())
    doc = render(data, tmp_path)
    raw = doc.path.read_bytes()
    # Watermark uses Helvetica-Bold at alpha 0.15 — both appear uncompressed in PDF metadata
    assert b"Helvetica-Bold" in raw
    assert b"/ca .15" in raw


# --- Data renderers: csv / json / jsonl ---

def test_csv_renderer_roundtrips(tmp_path: Path):
    import csv as _csv
    from recordforge.generators.data.customers import build_rows
    from recordforge.renderers.csv import render
    rows = build_rows(_fresh_rng(), count=8)
    doc = render("customers", rows, tmp_path)
    assert doc.format == "csv" and doc.path.suffix == ".csv"
    read = list(_csv.DictReader(doc.path.open(encoding="utf-8")))
    assert len(read) == 8
    assert read[0]["customer_id"] == "CUST-1000"


def test_json_renderer_is_valid_array(tmp_path: Path):
    import json as _json
    from recordforge.generators.data.customers import build_rows
    from recordforge.renderers.json import render
    rows = build_rows(_fresh_rng(), count=6)
    doc = render("customers", rows, tmp_path)
    data = _json.loads(doc.path.read_text(encoding="utf-8"))
    assert isinstance(data, list) and len(data) == 6


def test_jsonl_renderer_one_object_per_line(tmp_path: Path):
    import json as _json
    from recordforge.generators.data.customers import build_rows
    from recordforge.renderers.jsonl import render
    rows = build_rows(_fresh_rng(), count=7)
    doc = render("customers", rows, tmp_path)
    lines = [ln for ln in doc.path.read_text(encoding="utf-8").splitlines() if ln]
    assert len(lines) == 7
    assert all(isinstance(_json.loads(ln), dict) for ln in lines)


def test_messy_none_serializes_in_csv_and_json(tmp_path: Path):
    import json as _json
    from recordforge.generators.data.messy import build_rows
    from recordforge.renderers.csv import render as render_csv
    from recordforge.renderers.json import render as render_json
    rows = build_rows(_fresh_rng(), count=10)
    csv_doc = render_csv("messy", rows, tmp_path)
    json_doc = render_json("messy", rows, tmp_path)
    assert csv_doc.path.stat().st_size > 0
    assert len(_json.loads(json_doc.path.read_text(encoding="utf-8"))) == 10


# --- API: rows control and format validation ---

def test_generate_rows_controls_row_count(tmp_path: Path):
    import json as _json
    import recordforge as rf
    doc = rf.generate(type="customers", format="json", count=1, output=tmp_path, rows=17)[0]
    assert len(_json.loads(doc.path.read_text(encoding="utf-8"))) == 17


def test_generate_rejects_bad_data_format(tmp_path: Path):
    import recordforge as rf
    with pytest.raises(ValueError):
        rf.generate(type="customers", format="pdf", output=tmp_path)


def test_generate_rejects_data_format_on_documents(tmp_path: Path):
    import recordforge as rf
    with pytest.raises(ValueError):
        rf.generate(type="invoice", format="csv", output=tmp_path)


def test_generate_data_formats_are_reproducible(tmp_path: Path):
    import recordforge as rf
    a = rf.generate(type="customers", format="jsonl", output=tmp_path, seed=5, rows=12)[0]
    b = rf.generate(type="customers", format="jsonl", output=tmp_path, seed=5, rows=12)[0]
    assert a.path.read_text(encoding="utf-8") == b.path.read_text(encoding="utf-8")


# --- Desktop bridge seed ---

def test_ui_bridge_seed_is_reproducible(tmp_path: Path):
    from recordforge.ui.app import API

    def run(out):
        return API().generate({
            "mode": "data", "docTypes": ["customers", "payments"], "quantity": 2,
            "dataFormat": "csv", "rows": 6, "seed": 1234, "outputFolder": str(out),
        })

    a, b = run(tmp_path / "a"), run(tmp_path / "b")
    assert a["success"] and b["success"]

    def by_type(files):
        out = {}
        for f in files:
            out.setdefault(Path(f).name.split("_")[0], []).append(Path(f).read_text(encoding="utf-8"))
        return out

    ta, tb = by_type(a["files"]), by_type(b["files"])
    # same seed -> same content per type across runs
    assert sorted(ta) == sorted(tb)
    for key in ta:
        assert sorted(ta[key]) == sorted(tb[key])
    # two files of one type still differ (rng advances within a batch)
    assert ta["customers"][0] != ta["customers"][1]


# --- Checksum-valid identifiers ---

def _luhn_ok(number: str) -> bool:
    total = 0
    for i, ch in enumerate(reversed(number)):
        d = int(ch)
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return total % 10 == 0


def _iban_ok(iban: str) -> bool:
    rearranged = iban[4:] + iban[:4]
    numeric = "".join(str(ord(c) - 55) if c.isalpha() else c for c in rearranged)
    return int(numeric) % 97 == 1


def _aba_ok(n: str) -> bool:
    d = [int(x) for x in n]
    return (3 * (d[0] + d[3] + d[6]) + 7 * (d[1] + d[4] + d[7]) + (d[2] + d[5] + d[8])) % 10 == 0


def test_card_numbers_are_luhn_valid():
    from recordforge.core.faker_utils import rand_card
    rng = random.Random(3)
    cards = [rand_card(rng) for _ in range(200)]
    assert all(_luhn_ok(num) for _, num in cards)
    assert all(len(num) in (15, 16) for _, num in cards)


def test_ibans_are_mod97_valid():
    from recordforge.core.faker_utils import rand_iban
    rng = random.Random(3)
    assert all(_iban_ok(rand_iban(rng)) for _ in range(200))


def test_routing_numbers_valid_but_non_routable():
    from recordforge.core.faker_utils import rand_routing_number
    rng = random.Random(3)
    for _ in range(200):
        n = rand_routing_number(rng)
        assert len(n) == 9 and _aba_ok(n)
        assert n.startswith("99")  # unassigned Fed prefix -> non-routable


def test_payments_type_registered(tmp_path: Path):
    import recordforge as rf
    assert "payments" in rf.list_types()["data"]
    doc = rf.generate(type="payments", format="csv", rows=10, output=tmp_path)[0]
    assert doc.path.exists() and doc.path.stat().st_size > 0


# --- Edge-case corpus ---

def test_edge_cases_build_rows():
    from recordforge.generators.data.edge_cases import build_rows
    rows = build_rows(_fresh_rng(), count=60)
    assert len(rows) == 60
    assert {"edge_id", "category", "label", "value"} == set(rows[0].keys())
    # small count still spans several categories thanks to interleaving
    assert len({r["category"] for r in build_rows(_fresh_rng(), count=8)}) >= 5


def test_edge_cases_registered_and_json_valid(tmp_path: Path):
    import json as _json
    import recordforge as rf
    assert "edge_cases" in rf.list_types()["data"]
    doc = rf.generate(type="edge_cases", format="json", rows=200, output=tmp_path)[0]
    data = _json.loads(doc.path.read_text(encoding="utf-8"))  # must be valid JSON (no NaN/Infinity)
    assert len({r["category"] for r in data}) >= 8


def test_edge_cases_are_stable():
    from recordforge.generators.data.edge_cases import build_rows
    assert build_rows(random.Random(1), 30) == build_rows(random.Random(999), 30)


def test_edge_cases_xlsx_handles_control_chars(tmp_path: Path):
    """Regression: openpyxl rejects control chars; the renderer must escape them."""
    import openpyxl
    import recordforge as rf
    doc = rf.generate(type="edge_cases", format="xlsx", rows=200, seed=1, output=tmp_path)[0]
    ws = openpyxl.load_workbook(doc.path).active  # must open without IllegalCharacterError
    assert ws.max_row > 200


def test_xlsx_escapes_null_byte(tmp_path: Path):
    import openpyxl
    from recordforge.renderers.xlsx import render
    doc = render("t", [{"a": "x\x00y"}], tmp_path)
    ws = openpyxl.load_workbook(doc.path).active
    assert ws.cell(3, 1).value == "x\\x00y"


# --- Dirtying engine ---

def _clean_rows(n=20):
    return [{"id": i, "name": f"Name{i}", "amount": 100 + i} for i in range(n)]


def test_dirty_nulls_injects_none():
    from recordforge.core.dirty import apply_dirty
    out = apply_dirty(_clean_rows(), _fresh_rng(), {"nulls": 0.5})
    assert any(v is None for row in out for v in row.values())


def test_dirty_duplicates_grows_row_count():
    from recordforge.core.dirty import apply_dirty
    out = apply_dirty(_clean_rows(20), _fresh_rng(), {"duplicates": 1.0})
    assert len(out) == 40  # every row duplicated once at rate 1.0


def test_dirty_does_not_mutate_input():
    from recordforge.core.dirty import apply_dirty
    clean = _clean_rows(10)
    apply_dirty(clean, _fresh_rng(), {"nulls": 1.0})
    assert all(row["id"] is not None for row in clean)  # original untouched


def test_dirty_is_deterministic_under_seed():
    from recordforge.core.dirty import apply_dirty
    cfg = {"nulls": 0.2, "case_drift": 0.3, "encoding": 0.2, "duplicates": 0.1}
    a = apply_dirty(_clean_rows(), random.Random(9), cfg)
    b = apply_dirty(_clean_rows(), random.Random(9), cfg)
    assert a == b


def test_dirty_rejects_unknown_type():
    from recordforge.core.dirty import apply_dirty
    with pytest.raises(ValueError):
        apply_dirty(_clean_rows(), _fresh_rng(), {"bogus": 0.1})


def test_generate_applies_dirty(tmp_path: Path):
    import csv as _csv
    import recordforge as rf
    doc = rf.generate(
        type="customers", format="csv", rows=40, seed=3, output=tmp_path,
        dirty={"nulls": 0.3, "duplicates": 0.2},
    )[0]
    read = list(_csv.DictReader(doc.path.open(encoding="utf-8")))
    assert len(read) > 40  # duplicates added rows
    assert any(v == "" for row in read for v in row.values())  # nulls -> empty in CSV


def test_cli_parse_dirty():
    from recordforge.cli import _parse_dirty
    assert _parse_dirty("nulls=0.1, case_drift=0.2 ,duplicates=0.05") == {
        "nulls": 0.1, "case_drift": 0.2, "duplicates": 0.05,
    }
    with pytest.raises(ValueError):
        _parse_dirty("nulls")


# --- Relational bundle (referential integrity) ---

def test_build_related_has_no_orphan_foreign_keys():
    from recordforge.generators.related import build_related
    data = build_related(_fresh_rng(), {"customers": 30, "transactions": 120, "payments": 80})
    customer_ids = {c["customer_id"] for c in data["customers"]}
    txn_ids = {t["txn_id"] for t in data["transactions"]}
    assert len(data["customers"]) == 30
    assert len(data["transactions"]) == 120
    assert len(data["payments"]) == 80
    # every child key references a real parent
    assert all(t["customer_id"] in customer_ids for t in data["transactions"])
    assert all(p["customer_id"] in customer_ids for p in data["payments"])
    assert all(p["txn_id"] in txn_ids for p in data["payments"])


def test_build_related_payment_chain_is_consistent():
    from recordforge.generators.related import build_related
    data = build_related(_fresh_rng(), {"customers": 20, "transactions": 60, "payments": 40})
    txn_by_id = {t["txn_id"]: t for t in data["transactions"]}
    cust_by_id = {c["customer_id"]: c for c in data["customers"]}
    for pay in data["payments"]:
        txn = txn_by_id[pay["txn_id"]]
        # payment inherits the transaction's customer, amount, and currency
        assert pay["customer_id"] == txn["customer_id"]
        assert pay["amount"] == txn["amount"]
        assert pay["currency"] == txn["currency"]
        assert pay["account_holder"] == cust_by_id[pay["customer_id"]]["name"]


def test_build_related_defaults_and_clamps():
    from recordforge.generators.related import DEFAULT_COUNTS, build_related
    data = build_related(_fresh_rng())  # no spec -> defaults
    assert len(data["customers"]) == DEFAULT_COUNTS["customers"]
    clamped = build_related(_fresh_rng(), {"customers": 0, "transactions": -5, "payments": 3})
    assert len(clamped["customers"]) == 1  # counts clamp up to at least 1
    assert len(clamped["transactions"]) == 1


def test_generate_related_writes_three_datasets(tmp_path: Path):
    import csv as _csv
    import recordforge as rf
    docs = rf.generate_related(
        output=tmp_path, format="csv", seed=7,
        customers=25, transactions=90, payments=50,
    )
    assert len(docs) == 3
    by_type = {d.doc_type: d for d in docs}
    assert set(by_type) == {"customers", "transactions", "payments"}
    for doc in docs:
        assert doc.path.exists() and doc.path.stat().st_size > 0
    # re-load and re-check FK integrity end to end through the CSV renderer
    customer_ids = {r["customer_id"] for r in _csv.DictReader(by_type["customers"].path.open(encoding="utf-8"))}
    txns = list(_csv.DictReader(by_type["transactions"].path.open(encoding="utf-8")))
    assert all(t["customer_id"] in customer_ids for t in txns)


def test_generate_related_is_reproducible(tmp_path: Path):
    import recordforge as rf
    a = rf.generate_related(output=tmp_path / "a", format="jsonl", seed=11, customers=10, transactions=40, payments=20)
    b = rf.generate_related(output=tmp_path / "b", format="jsonl", seed=11, customers=10, transactions=40, payments=20)
    a_by = {d.doc_type: d.path.read_text(encoding="utf-8") for d in a}
    b_by = {d.doc_type: d.path.read_text(encoding="utf-8") for d in b}
    assert a_by == b_by


def test_generate_related_rejects_bad_format(tmp_path: Path):
    import recordforge as rf
    with pytest.raises(ValueError):
        rf.generate_related(output=tmp_path, format="pdf")


def test_cli_generate_related(tmp_path: Path):
    from typer.testing import CliRunner
    from recordforge.cli import app
    result = CliRunner().invoke(app, [
        "generate-related", "--format", "json", "--output", str(tmp_path),
        "--customers", "15", "--transactions", "40", "--payments", "20", "--seed", "2",
    ])
    assert result.exit_code == 0
    assert "Generated 3 linked dataset(s)." in result.stdout
    assert len(list(tmp_path.glob("*.json"))) == 3


# --- Schema-driven generation (C2) ---

EXAMPLES = Path(__file__).resolve().parent.parent / "examples"


def _write_schema(tmp_path: Path, name: str, body: str) -> Path:
    p = tmp_path / name
    p.write_text(body, encoding="utf-8")
    return p


def test_schema_topo_order_parent_before_child():
    from recordforge.schema import Column, Dataset, Schema, topo_order
    # child declared before parent on purpose
    child = Dataset("orders", 5, [
        Column("order_id", "id"),
        Column("customer_id", "fk", {"ref": "customers.customer_id"}),
    ])
    parent = Dataset("customers", 5, [Column("customer_id", "id")])
    order = topo_order(Schema("s", [child, parent]))
    assert order.index("customers") < order.index("orders")


def test_schema_cross_table_cycle_is_rejected():
    from recordforge.schema import Column, Dataset, Schema, topo_order
    a = Dataset("a", 3, [Column("a_id", "id"), Column("b_id", "fk", {"ref": "b.b_id"})])
    b = Dataset("b", 3, [Column("b_id", "id"), Column("a_id", "fk", {"ref": "a.a_id"})])
    with pytest.raises(ValueError, match="circular"):
        topo_order(Schema("s", [a, b]))


def test_schema_generate_has_no_orphan_fks(tmp_path: Path):
    from recordforge.schema import generate_datasets, load_schema
    data = generate_datasets(_fresh_rng(), load_schema(EXAMPLES / "shop.yml"))
    customer_ids = {r["customer_id"] for r in data["customers"]}
    order_ids = {r["order_id"] for r in data["orders"]}
    assert all(o["customer_id"] in customer_ids for o in data["orders"])
    assert all(p["order_id"] in order_ids for p in data["order_payments"])
    assert all(p["customer_id"] in customer_ids for p in data["order_payments"])


def test_schema_self_reference_builds_acyclic_hierarchy():
    from recordforge.schema import generate_datasets, load_schema
    data = generate_datasets(_fresh_rng(), load_schema(EXAMPLES / "org.json"))
    emps = data["employees"]
    ids = {e["employee_id"] for e in emps}
    # first employee is a null root; nobody manages themselves; managers exist
    assert emps[0]["manager_id"] is None
    for e in emps:
        assert e["manager_id"] != e["employee_id"]
        if e["manager_id"] is not None:
            assert e["manager_id"] in ids
    assert any(e["manager_id"] is not None for e in emps)


def test_schema_generate_schema_writes_files_and_is_reproducible(tmp_path: Path):
    import recordforge as rf
    a = rf.generate_schema(EXAMPLES / "shop.yml", output=tmp_path / "a", format="json", seed=4)
    b = rf.generate_schema(EXAMPLES / "shop.yml", output=tmp_path / "b", format="json", seed=4)
    assert {d.doc_type for d in a} == {"customers", "orders", "order_payments"}
    a_by = {d.doc_type: d.path.read_text(encoding="utf-8") for d in a}
    b_by = {d.doc_type: d.path.read_text(encoding="utf-8") for d in b}
    assert a_by == b_by


def test_schema_decimal_column_serializes_json(tmp_path: Path):
    import json as _json
    import recordforge as rf
    docs = rf.generate_schema(EXAMPLES / "shop.yml", output=tmp_path, format="json", seed=1)
    orders = next(d for d in docs if d.doc_type == "orders")
    rows = _json.loads(orders.path.read_text(encoding="utf-8"))  # must parse
    assert all("." in str(r["amount"]) for r in rows)  # decimal rendered with places


def test_schema_rejects_unknown_type(tmp_path: Path):
    from recordforge.schema import load_schema
    p = _write_schema(tmp_path, "bad.json",
        '{"datasets":[{"name":"t","count":2,"columns":[{"name":"x","type":"bogus"}]}]}')
    with pytest.raises(ValueError, match="unknown type"):
        load_schema(p)


def test_schema_rejects_fk_to_missing_table(tmp_path: Path):
    from recordforge.schema import load_schema
    p = _write_schema(tmp_path, "bad.json",
        '{"datasets":[{"name":"t","count":2,"columns":['
        '{"name":"id","type":"id"},{"name":"fk","type":"fk","ref":"ghost.id"}]}]}')
    with pytest.raises(ValueError, match="unknown table"):
        load_schema(p)


def test_schema_rejects_choice_without_values(tmp_path: Path):
    from recordforge.schema import load_schema
    p = _write_schema(tmp_path, "bad.json",
        '{"datasets":[{"name":"t","count":2,"columns":[{"name":"s","type":"choice"}]}]}')
    with pytest.raises(ValueError, match="choice"):
        load_schema(p)


def test_cli_generate_schema(tmp_path: Path):
    from typer.testing import CliRunner
    from recordforge.cli import app
    result = CliRunner().invoke(app, [
        "generate-schema", "--schema", str(EXAMPLES / "shop.yml"),
        "--format", "csv", "--output", str(tmp_path), "--seed", "3",
    ])
    assert result.exit_code == 0
    assert "Generated 3 dataset(s) from schema." in result.stdout
    assert len(list(tmp_path.glob("*.csv"))) == 3


# --- SQL renderer (C3) ---

def test_sql_literal_and_ident_escaping():
    from decimal import Decimal
    from recordforge.renderers.sql import _ident, _literal
    assert _literal(None) == "NULL"
    assert _literal(True) == "TRUE" and _literal(False) == "FALSE"
    assert _literal(42) == "42"
    assert _literal(Decimal("19.95")) == "19.95"
    assert _literal("O'Brien") == "'O''Brien'"   # single quote doubled
    assert _ident('weird"name') == '"weird""name"'  # double quote doubled


def test_sql_single_table_render(tmp_path: Path):
    from recordforge.generators.data.customers import build_rows
    from recordforge.renderers.sql import render
    rows = build_rows(_fresh_rng(), count=5)
    doc = render("customers", rows, tmp_path)
    text = doc.path.read_text(encoding="utf-8")
    assert doc.format == "sql" and doc.path.suffix == ".sql"
    assert 'INSERT INTO "customers"' in text
    assert text.count("(") >= 5  # at least one value tuple per row


def test_generate_sql_format(tmp_path: Path):
    import recordforge as rf
    doc = rf.generate(type="payments", format="sql", rows=8, seed=1, output=tmp_path)[0]
    assert doc.path.suffix == ".sql"
    assert 'INSERT INTO "payments"' in doc.path.read_text(encoding="utf-8")


def test_related_sql_bundle_is_single_fk_ordered_file(tmp_path: Path):
    import recordforge as rf
    docs = rf.generate_related(output=tmp_path, format="sql", seed=1,
                               customers=20, transactions=60, payments=40)
    assert len(docs) == 1  # one combined file, not three
    text = docs[0].path.read_text(encoding="utf-8")
    # parents inserted before children
    assert text.index('INSERT INTO "customers"') < text.index('INSERT INTO "transactions"')
    assert text.index('INSERT INTO "transactions"') < text.index('INSERT INTO "payments"')


def test_related_sql_bundle_executes_in_sqlite(tmp_path: Path):
    """The emitted INSERTs load and join cleanly in a real database."""
    import re
    import sqlite3
    import recordforge as rf
    doc = rf.generate_related(output=tmp_path, format="sql", seed=2,
                              customers=25, transactions=80, payments=50)[0]
    sql = doc.path.read_text(encoding="utf-8")
    # Build CREATE TABLE from each INSERT's column list (all TEXT is fine here).
    cols_by_table: dict[str, str] = {}
    for m in re.finditer(r'INSERT INTO "(\w+)" \(([^)]+)\) VALUES', sql):
        cols_by_table.setdefault(m.group(1), m.group(2))
    ddl = "".join(
        f'CREATE TABLE "{t}" ({", ".join(c.strip() + " TEXT" for c in cols.split(","))});\n'
        for t, cols in cols_by_table.items()
    )
    con = sqlite3.connect(":memory:")
    con.executescript(ddl + sql)  # raises if the SQL is malformed
    cur = con.cursor()
    customers = {r[0] for r in cur.execute('SELECT "customer_id" FROM "customers"')}
    txn_cust = {r[0] for r in cur.execute('SELECT "customer_id" FROM "transactions"')}
    pay_txn = {r[0] for r in cur.execute('SELECT "txn_id" FROM "payments"')}
    txns = {r[0] for r in cur.execute('SELECT "txn_id" FROM "transactions"')}
    assert txn_cust <= customers      # no orphan transactions
    assert pay_txn <= txns            # no orphan payments
    con.close()


def test_schema_sql_bundle_is_topologically_ordered(tmp_path: Path):
    """Even when a child dataset is declared before its parent, SQL is FK-ordered."""
    import recordforge as rf
    p = _write_schema(tmp_path, "rev.json",
        '{"datasets":['
        '{"name":"orders","count":5,"columns":['
        '{"name":"order_id","type":"id"},'
        '{"name":"customer_id","type":"fk","ref":"customers.customer_id"}]},'
        '{"name":"customers","count":5,"columns":[{"name":"customer_id","type":"id"}]}'
        ']}')
    doc = rf.generate_schema(p, output=tmp_path, format="sql", seed=1)[0]
    text = doc.path.read_text(encoding="utf-8")
    assert text.index('INSERT INTO "customers"') < text.index('INSERT INTO "orders"')


def test_sql_bundle_is_reproducible(tmp_path: Path):
    import recordforge as rf
    a = rf.generate_related(output=tmp_path / "a", format="sql", seed=9,
                            customers=10, transactions=30, payments=15)[0]
    b = rf.generate_related(output=tmp_path / "b", format="sql", seed=9,
                            customers=10, transactions=30, payments=15)[0]
    assert a.path.read_text(encoding="utf-8") == b.path.read_text(encoding="utf-8")


# --- Parquet renderer (C3, optional pyarrow) ---

def test_parquet_round_trips(tmp_path: Path):
    pytest.importorskip("pyarrow")
    import pyarrow.parquet as pq
    import recordforge as rf
    doc = rf.generate(type="customers", format="parquet", rows=30, seed=1, output=tmp_path)[0]
    assert doc.path.suffix == ".parquet"
    table = pq.read_table(doc.path)
    assert table.num_rows == 30
    assert "customer_id" in table.column_names


def test_parquet_handles_mixed_type_edge_cases(tmp_path: Path):
    pytest.importorskip("pyarrow")
    import pyarrow.parquet as pq
    import recordforge as rf
    # edge_cases 'value' column mixes ints, floats, strings, and nulls
    doc = rf.generate(type="edge_cases", format="parquet", rows=200, seed=1, output=tmp_path)[0]
    table = pq.read_table(doc.path)  # must read back without error
    assert table.num_rows == 200
    # the mixed-type 'value' column falls back to text rather than crashing
    assert table.schema.field("value").type == "string"


def test_related_parquet_keeps_fk_integrity(tmp_path: Path):
    pytest.importorskip("pyarrow")
    import pyarrow.parquet as pq
    import recordforge as rf
    docs = rf.generate_related(output=tmp_path, format="parquet", seed=1,
                               customers=20, transactions=60, payments=40)
    assert len(docs) == 3 and all(d.path.suffix == ".parquet" for d in docs)
    by = {d.doc_type: pq.read_table(d.path).to_pylist() for d in docs}
    customer_ids = {r["customer_id"] for r in by["customers"]}
    assert all(t["customer_id"] in customer_ids for t in by["transactions"])
    assert all(p["customer_id"] in customer_ids for p in by["payments"])


# --- Seed reproducibility ---

def test_same_seed_same_doc_number():
    from recordforge.generators.documents.invoice import build
    set_seed(99)
    a = build(get_rng())
    set_seed(99)
    b = build(get_rng())
    assert a.doc_number == b.doc_number


def test_same_seed_same_total_due():
    from recordforge.generators.documents.invoice import build
    set_seed(42)
    a = build(get_rng())
    set_seed(42)
    b = build(get_rng())
    assert a.total_due == b.total_due
